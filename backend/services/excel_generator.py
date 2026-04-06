"""Generadores de Excel para Matrices"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging
from models.matrices import MatrizSST, MatrizLegal, NivelRiesgo

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Genera archivos Excel profesionales para matrices"""
    
    @staticmethod
    def generar_matriz_sst(matriz: MatrizSST) -> bytes:
        """Genera Excel para Matriz SST (GTC 45)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz GTC 45"
        
        # Estilos
        header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # TÍTULO
        ws.merge_cells('A1:Q1')
        ws['A1'] = f"MATRIZ DE IDENTIFICACIÓN DE PELIGROS Y VALORACIÓN DE RIESGOS - GTC 45"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # INFO EMPRESA
        ws.merge_cells('A2:Q2')
        ws['A2'] = f"EMPRESA: {matriz.empresa} | DOCUMENTO: {matriz.documento_origen} | FECHA: {matriz.created_at.strftime('%Y-%m-%d')}"
        ws['A2'].font = Font(size=10)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # ESTADÍSTICAS
        ws.merge_cells('A3:Q3')
        ws['A3'] = f"Total Riesgos: {matriz.total_riesgos} | Críticos: {matriz.riesgos_criticos} | Altos: {matriz.riesgos_altos} | Medios: {matriz.riesgos_medios} | Bajos: {matriz.riesgos_bajos}"
        ws['A3'].font = Font(size=10, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # HEADERS
        headers = [
            "ID", "Proceso", "Zona/Lugar", "Actividad", "Clasificación Peligro",
            "Descripción Peligro", "Efectos Posibles", "Controles Existentes",
            "ND", "NE", "NP", "Probabilidad", "NC", "NR", "Nivel Riesgo",
            "Controles Propuestos", "Fuente"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # DATOS
        row_num = 6
        for riesgo in matriz.riesgos:
            # Colorear según nivel de riesgo
            if riesgo.interpretacion_riesgo == NivelRiesgo.CRITICO:
                row_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                font_color = "FFFFFF"
            elif riesgo.interpretacion_riesgo == NivelRiesgo.ALTO:
                row_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
                font_color = "FFFFFF"
            elif riesgo.interpretacion_riesgo == NivelRiesgo.MEDIO:
                row_fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
                font_color = "000000"
            else:
                row_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
                font_color = "FFFFFF"
            
            data = [
                riesgo.id_riesgo,
                riesgo.proceso,
                riesgo.zona_lugar,
                riesgo.actividad,
                riesgo.peligro.clasificacion,
                riesgo.peligro.descripcion,
                ", ".join(riesgo.peligro.efectos_posibles),
                ", ".join(riesgo.controles_existentes) if riesgo.controles_existentes else "N/A",
                riesgo.nivel_deficiencia,
                riesgo.nivel_exposicion,
                riesgo.nivel_probabilidad,
                riesgo.interpretacion_probabilidad,
                riesgo.nivel_consecuencia,
                riesgo.nivel_riesgo,
                riesgo.interpretacion_riesgo.value,
                ", ".join(riesgo.controles_propuestos),
                riesgo.peligro.fuente
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if col_num == 15:  # Nivel de Riesgo
                    cell.fill = row_fill
                    cell.font = Font(bold=True, color=font_color)
            
            row_num += 1
        
        # Ajustar anchos
        column_widths = [10, 15, 15, 20, 15, 25, 20, 20, 6, 6, 6, 12, 6, 6, 12, 25, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Altura de filas
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[5].height = 40
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"✅ Excel SST generado: {len(matriz.riesgos)} filas")
        return output.getvalue()
    
    @staticmethod
    def generar_matriz_legal(matriz: MatrizLegal) -> bytes:
        """Genera Excel para Matriz de Riesgos Legales"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Riesgos Legales"
        
        # Estilos
        header_fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # TÍTULO
        ws.merge_cells('A1:O1')
        ws['A1'] = f"MATRIZ DE RIESGOS LEGALES"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # INFO EMPRESA
        ws.merge_cells('A2:O2')
        ws['A2'] = f"EMPRESA: {matriz.empresa} | DOCUMENTO: {matriz.documento_origen} | FECHA: {matriz.created_at.strftime('%Y-%m-%d')}"
        ws['A2'].font = Font(size=10)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # ESTADÍSTICAS
        ws.merge_cells('A3:O3')
        ws['A3'] = f"Total Riesgos: {matriz.total_riesgos} | Críticos: {matriz.riesgos_criticos} | Altos: {matriz.riesgos_altos} | Medios: {matriz.riesgos_medios} | Bajos: {matriz.riesgos_bajos}"
        ws['A3'].font = Font(size=10, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # HEADERS
        headers = [
            "ID", "Categoría", "Descripción", "Normativa Aplicable",
            "Cláusulas Relevantes", "Probabilidad", "Imp. Financiero",
            "Imp. Reputacional", "Imp. Operacional", "NR Calc", "Nivel Riesgo",
            "Controles Actuales", "Acciones Mitigación", "Responsable", "Fuente"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # DATOS
        row_num = 6
        for riesgo in matriz.riesgos:
            # Colorear según nivel de riesgo
            if riesgo.nivel_riesgo == NivelRiesgo.CRITICO:
                row_fill = PatternFill(start_color="7F1D1D", end_color="7F1D1D", fill_type="solid")
                font_color = "FFFFFF"
            elif riesgo.nivel_riesgo == NivelRiesgo.ALTO:
                row_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                font_color = "FFFFFF"
            elif riesgo.nivel_riesgo == NivelRiesgo.MEDIO:
                row_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                font_color = "000000"
            else:
                row_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
                font_color = "FFFFFF"
            
            data = [
                riesgo.id_riesgo,
                riesgo.categoria.value,
                riesgo.descripcion,
                ", ".join(riesgo.normativa_aplicable),
                ", ".join(riesgo.clausulas_relevantes),
                riesgo.probabilidad_ocurrencia,
                riesgo.impacto_financiero,
                riesgo.impacto_reputacional,
                riesgo.impacto_operacional,
                round(riesgo.nivel_riesgo_calculado, 2),
                riesgo.nivel_riesgo.value,
                ", ".join(riesgo.controles_actuales) if riesgo.controles_actuales else "N/A",
                ", ".join(riesgo.acciones_mitigacion),
                riesgo.responsable_sugerido,
                riesgo.fuente_documento
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if col_num == 11:  # Nivel de Riesgo
                    cell.fill = row_fill
                    cell.font = Font(bold=True, color=font_color)
            
            row_num += 1
        
        # Ajustar anchos
        column_widths = [10, 15, 30, 20, 15, 10, 12, 12, 12, 8, 12, 20, 25, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Altura de filas
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[5].height = 40
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"✅ Excel Legal generado: {len(matriz.riesgos)} filas")
        return output.getvalue()

# Instancia singleton
excel_generator = ExcelGenerator()
