from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

from models.gtc45_models import MatrizGTC45
from services.gtc45_calculator import obtener_color_riesgo

def generar_excel_gtc45(matriz: MatrizGTC45) -> bytes:
    """Genera archivo Excel con formato GTC 45:2012"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz GTC 45"
    
    # Estilos
    header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:T1')
    title_cell = ws['A1']
    title_cell.value = f"MATRIZ DE IDENTIFICACIÓN DE PELIGROS, EVALUACIÓN Y VALORACIÓN DE RIESGOS - GTC 45:2012"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Información de la empresa
    ws.merge_cells('A2:T2')
    ws['A2'].value = f"EMPRESA: {matriz.nombre_empresa}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:T3')
    ws['A3'].value = f"Fecha de elaboración: {datetime.fromisoformat(matriz.fecha_elaboracion).strftime('%d/%m/%Y')} | Responsable: {matriz.responsable_elaboracion}"
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    
    # Encabezados GTC 45
    headers = [
        'Proceso',
        'Actividad',
        'Tipo',
        'Tarea',
        'Clasificación Peligro',
        'Descripción Peligro',
        'Fuente',
        'Efectos Posibles',
        'ND',
        'Nivel ND',
        'NE',
        'Nivel NE',
        'NP',
        'Nivel NP',
        'NC',
        'Nivel NC',
        'NR',
        'Interp.',
        'Aceptabilidad',
        'Controles Existentes',
        'Controles Recomendados'
    ]
    
    ws.append(headers)
    header_row = ws.max_row
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    for riesgo in matriz.riesgos:
        val = riesgo.valoracion
        
        controles_exist = "\n".join([
            f"{ctrl.jerarquia}: {ctrl.descripcion}" 
            for ctrl in riesgo.controles_existentes
        ]) if riesgo.controles_existentes else "Ninguno"
        
        controles_recom = "\n".join([
            f"{ctrl.jerarquia}: {ctrl.descripcion}" 
            for ctrl in riesgo.controles_recomendados
        ]) if riesgo.controles_recomendados else "Ninguno"
        
        efectos = "\n".join(riesgo.peligro.efectos_posibles) if riesgo.peligro.efectos_posibles else "N/A"
        
        row_data = [
            riesgo.proceso.nombre,
            riesgo.actividad.nombre,
            riesgo.actividad.tipo,
            riesgo.tarea.nombre,
            riesgo.peligro.clasificacion,
            riesgo.peligro.descripcion,
            riesgo.peligro.fuente or "N/A",
            efectos,
            val.nd_valor,
            val.nd_nivel,
            val.ne_valor,
            val.ne_nivel,
            val.np_valor,
            val.np_nivel,
            val.nc_valor,
            val.nc_nivel,
            val.nr_valor,
            val.interpretacion,
            val.aceptabilidad,
            controles_exist,
            controles_recom
        ]
        
        ws.append(row_data)
        current_row = ws.max_row
        
        # Aplicar estilos
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Color semáforo para NR y Interpretación
        color_hex = obtener_color_riesgo(val.interpretacion).replace('#', '')
        nr_cell = ws.cell(row=current_row, column=17)  # NR
        interp_cell = ws.cell(row=current_row, column=18)  # Interpretación
        
        nr_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        interp_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
        if val.interpretacion in ["I", "II"]:
            nr_cell.font = Font(bold=True, color="FFFFFF")
            interp_cell.font = Font(bold=True, color="FFFFFF")
        else:
            nr_cell.font = Font(bold=True)
            interp_cell.font = Font(bold=True)
    
    # Ajustar anchos de columna
    column_widths = [15, 20, 12, 25, 18, 30, 20, 30, 6, 12, 6, 12, 6, 12, 6, 15, 8, 8, 18, 35, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[header_row].height = 40
    
    # Hoja de resumen
    resumen_sheet = wb.create_sheet("Resumen Ejecutivo")
    resumen_sheet.merge_cells('A1:F1')
    resumen_title = resumen_sheet['A1']
    resumen_title.value = "RESUMEN EJECUTIVO"
    resumen_title.font = Font(bold=True, size=14)
    resumen_title.alignment = Alignment(horizontal='center')
    
    resumen_sheet.append([])
    resumen_sheet.merge_cells('A3:F10')
    resumen_cell = resumen_sheet['A3']
    resumen_cell.value = matriz.resumen_ejecutivo
    resumen_cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Estadísticas
    stats = matriz.estadisticas
    resumen_sheet.append([])
    resumen_sheet.append([])
    resumen_sheet.append(['ESTADÍSTICAS'])
    resumen_sheet['A13'].font = Font(bold=True, size=12)
    
    resumen_sheet.append([])
    resumen_sheet.append(['Total de Riesgos Identificados:', stats['total_riesgos']])
    resumen_sheet.append([])
    resumen_sheet.append(['Por Nivel de Riesgo:'])
    resumen_sheet['A17'].font = Font(bold=True)
    
    resumen_sheet.append(['  Nivel I (Crítico - Rojo):', stats['criticos']])
    resumen_sheet.append(['  Nivel II (Alto - Naranja):', stats['altos']])
    resumen_sheet.append(['  Nivel III (Medio - Amarillo):', stats['medios']])
    resumen_sheet.append(['  Nivel IV (Bajo - Verde):', stats['bajos']])
    
    # Colores en estadísticas
    resumen_sheet['A18'].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    resumen_sheet['A18'].font = Font(color="FFFFFF", bold=True)
    resumen_sheet['A19'].fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    resumen_sheet['A19'].font = Font(color="FFFFFF", bold=True)
    resumen_sheet['A20'].fill = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid")
    resumen_sheet['A20'].font = Font(bold=True)
    resumen_sheet['A21'].fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    resumen_sheet['A21'].font = Font(color="FFFFFF", bold=True)
    
    resumen_sheet.append([])
    resumen_sheet.append(['Por Clasificación de Peligro:'])
    resumen_sheet['A23'].font = Font(bold=True)
    
    for clasificacion, cantidad in stats['por_clasificacion'].items():
        resumen_sheet.append([f'  {clasificacion}:', cantidad])
    
    resumen_sheet.column_dimensions['A'].width = 50
    resumen_sheet.column_dimensions['B'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()